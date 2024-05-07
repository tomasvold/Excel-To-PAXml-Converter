import tkinter as tk
from tkinter import filedialog
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from datetime import datetime
import os  # Import os module for file path handling

def excel_to_xml(file_path, output_path):
    # Load Excel workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Create root element in XML
    paxml = ET.Element("paxml")
    paxml.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
    paxml.set("xsi:noNamespaceSchemaLocation", "http://www.paxml.se/2.0/paxml.xsd")
    paxml.text = '\n'  

    # Create header element and its children
    header = ET.SubElement(paxml, "header")
    header.text = '\n\t'  
    format_element = ET.SubElement(header, "format")
    format_element.text = "LÖNIN"
    format_element.tail = '\n\t' 
    version = ET.SubElement(header, "version")
    version.text = '2.0'
    version.tail ='\n'
    header.tail = '\n'

    # Create schematransaktioner element
    schematransaktioner = ET.SubElement(paxml, "schematransaktioner")
    schematransaktioner.text = '\n\t'
    schematransaktioner.tail = '\n'

    # Create a dictionary to store schema elements
    schema_dict = {}

    # Iterate through rows in excel
    for row in ws.iter_rows(min_row=2, values_only=True):
        anstid, datum, timmar = row

        # Convert datum to yyyy-MM-dd format
        datum_formatted = datum.strftime("%Y-%m-%d")

        # Format timmar to two decimal places
        timmar_formatted = "{:.2f}".format(timmar)

        # Create schema element if not exists
        if anstid not in schema_dict:
            schema_dict[anstid] = ET.SubElement(schematransaktioner, "schema")
            schema_dict[anstid].set("anstid", str(anstid))
            schema_dict[anstid].text = '\n\t\t' if schema_dict else ''

        # Create dag element under schema
        dag = ET.SubElement(schema_dict[anstid], "dag")
        dag.set("datum", datum_formatted)
        dag.set("timmar", str(timmar_formatted))
        dag.tail = '\n\t\t'  

    # Add newline after each schema element
    for schema in schema_dict.values():
        schema.tail = '\n\t'  

    # Create XML tree
    tree = ET.ElementTree(paxml)
    tree.write(output_path, encoding="utf-8", xml_declaration=True)

def browse_file():
    # Create the main Tkinter GUI window
    root = tk.Tk()
    root.title("Excel to PAXml 2.0 Converter")

    # Set window size
    window_width = 600
    window_height = 120
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_coordinate = (screen_width / 2) - (window_width / 2)
    y_coordinate = (screen_height / 2) - (window_height / 2)
    root.geometry("%dx%d+%d+%d" % (window_width, window_height, x_coordinate, y_coordinate))

    # Function to trigger file dialog
    def open_file_dialog():
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if file_path:
                # Determine output path for XML file
                output_path = os.path.join(os.path.dirname(file_path), "PAXml-Schema.xml")
                
                # Convert Excel to XML
                excel_to_xml(file_path, output_path)
                status_label.config(text="Konvertering lyckades! PAXml-Schema.xml är nu nerladdad - du kan nu stänga programmet")
        except Exception as e:
            status_label.config(text="Ogiltigt filformat. Testa med en annan fil.\nAccepterade filformat: .xlsx med följande kolumner:\n\nanstid, datum, timmar")

    # Create a label to display the conversion status
    status_label = tk.Label(root, text="Välj en fil att konvertera", padx=10, pady=10)
    status_label.pack()

    # Create a button to trigger the file dialog
    browse_button = tk.Button(root, text="Browse", command=open_file_dialog)
    browse_button.pack()

    # Run the Tkinter event loop
    root.mainloop()

if __name__ == "__main__":
    # Call browse_file() only if the script is executed directly
    browse_file()

#to compile: python -m PyInstaller --onedir -w ExcelPAXml.py 
#to create the wizard, use Inno setup compiler
